import os, uuid
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from openpyxl import load_workbook, Workbook
import psycopg
from psycopg.rows import dict_row

BASE_DIR = Path(__file__).resolve().parent
UPLOADS = BASE_DIR / "uploads"
UPLOADS.mkdir(exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret")

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL belum diatur di Render.")

def db():
    return psycopg.connect(DATABASE_URL, row_factory=dict_row)

def init_db():
    with db() as c:
        with c.cursor() as cur:
            cur.execute("""
            CREATE TABLE IF NOT EXISTS settings(
              id INTEGER PRIMARY KEY, language TEXT DEFAULT 'id', theme TEXT DEFAULT 'green'
            );
            INSERT INTO settings(id,language,theme) VALUES(1,'id','green')
            ON CONFLICT(id) DO NOTHING;

            CREATE TABLE IF NOT EXISTS stores(
              id BIGSERIAL PRIMARY KEY, code TEXT UNIQUE NOT NULL, name TEXT NOT NULL, active BOOLEAN DEFAULT TRUE
            );
            CREATE TABLE IF NOT EXISTS products(
              id BIGSERIAL PRIMARY KEY, sku TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
              category TEXT DEFAULT '', unit TEXT DEFAULT 'PCS', buy_price NUMERIC DEFAULT 0,
              supplier TEXT DEFAULT '', stock NUMERIC DEFAULT 0, minimum NUMERIC DEFAULT 0, active BOOLEAN DEFAULT TRUE
            );
            CREATE TABLE IF NOT EXISTS inbounds(
              id BIGSERIAL PRIMARY KEY, number TEXT UNIQUE NOT NULL, arrival_date DATE,
              supplier_code TEXT, supplier_name TEXT, sku TEXT, exp_date DATE,
              product_name TEXT, price NUMERIC DEFAULT 0, qty NUMERIC DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS requests(
              id BIGSERIAL PRIMARY KEY, number TEXT UNIQUE NOT NULL, store_id BIGINT REFERENCES stores(id),
              request_date DATE, status TEXT DEFAULT 'OPEN', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS request_items(
              id BIGSERIAL PRIMARY KEY, request_id BIGINT REFERENCES requests(id) ON DELETE CASCADE,
              sku TEXT, product_name TEXT, qty NUMERIC DEFAULT 0, fulfilled NUMERIC DEFAULT 0, shortage NUMERIC DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS outbounds(
              id BIGSERIAL PRIMARY KEY, number TEXT UNIQUE NOT NULL, request_id BIGINT REFERENCES requests(id),
              store_id BIGINT REFERENCES stores(id), outbound_date DATE, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS outbound_items(
              id BIGSERIAL PRIMARY KEY, outbound_id BIGINT REFERENCES outbounds(id) ON DELETE CASCADE,
              sku TEXT, qty NUMERIC DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS logs(
              id BIGSERIAL PRIMARY KEY, action TEXT, detail TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cur.execute("SELECT COUNT(*) AS n FROM stores")
            if cur.fetchone()["n"] == 0:
                for i in range(1,17):
                    cur.execute("INSERT INTO stores(code,name) VALUES(%s,%s) ON CONFLICT(code) DO NOTHING",
                                (f"T{i:03d}", f"Toko {i:02d}"))

def log(action, detail=""):
    with db() as c:
        with c.cursor() as cur:
            cur.execute("INSERT INTO logs(action,detail) VALUES(%s,%s)", (action, detail))

def next_no(prefix, table):
    today = datetime.now().strftime("%Y%m%d")
    with db() as c:
        with c.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE number LIKE %s", (f"{prefix}-{today}-%",))
            n = cur.fetchone()["n"] + 1
    return f"{prefix}-{today}-{n:04d}"

@app.before_request
def startup():
    if not getattr(app, "_db_ready", False):
        init_db()
        app._db_ready = True

@app.context_processor
def common():
    with db() as c:
        with c.cursor() as cur:
            cur.execute("SELECT * FROM settings WHERE id=1")
            s = cur.fetchone()
    return dict(settings=s, now=datetime.now())

@app.route("/")
def dashboard():
    with db() as c:
        with c.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS n FROM stores WHERE active=TRUE"); stores=cur.fetchone()["n"]
            cur.execute("SELECT COUNT(*) AS n FROM products WHERE active=TRUE"); products=cur.fetchone()["n"]
            cur.execute("SELECT COUNT(*) AS n FROM requests"); requests=cur.fetchone()["n"]
            cur.execute("SELECT COUNT(*) AS n FROM outbounds"); outbounds=cur.fetchone()["n"]
            cur.execute("SELECT * FROM products WHERE active=TRUE AND stock<=minimum ORDER BY stock ASC LIMIT 10"); low=cur.fetchall()
    return render_template("dashboard.html", data={"stores":stores,"products":products,"requests":requests,"outbounds":outbounds}, low=low)

@app.route("/stores", methods=["GET","POST"])
def stores():
    q=request.args.get("q","").strip()
    if request.method=="POST":
        try:
            with db() as c:
                with c.cursor() as cur:
                    cur.execute("INSERT INTO stores(code,name) VALUES(%s,%s)",(request.form["code"].strip().upper(),request.form["name"].strip()))
            log("Tambah Toko", f'{request.form["code"]} - {request.form["name"]}')
            flash("Toko berhasil ditambahkan.","ok")
        except Exception as e: flash("Gagal: "+str(e),"err")
        return redirect(url_for("stores"))
    with db() as c:
        with c.cursor() as cur:
            cur.execute("SELECT * FROM stores WHERE code ILIKE %s OR name ILIKE %s ORDER BY id",(f"%{q}%",f"%{q}%")); rows=cur.fetchall()
    return render_template("stores.html", stores=rows, q=q)

@app.post("/stores/<int:id>/toggle")
def toggle_store(id):
    with db() as c:
        with c.cursor() as cur:
            cur.execute("UPDATE stores SET active=NOT active WHERE id=%s RETURNING name,active",(id,)); r=cur.fetchone()
    log("Ubah Status Toko",r["name"]); return redirect(url_for("stores"))

@app.route("/products", methods=["GET","POST"])
def products():
    if request.method=="POST":
        f=request.files.get("file")
        if not f or not f.filename.lower().endswith((".xlsx",".xlsm")):
            flash("Pilih file Excel .xlsx/.xlsm.","err"); return redirect(url_for("products"))
        path=UPLOADS/(str(uuid.uuid4())+"_"+f.filename); f.save(path)
        try:
            ws=load_workbook(path,data_only=True).active
            headers=[str(x.value).strip().lower() if x.value is not None else "" for x in ws[1]]
            aliases={"sku":["sku"],"name":["nama barang","nama","name"],"category":["kategori","category"],"unit":["satuan","unit"],"buy_price":["harga pembelian","buy price","harga"],"supplier":["supplier","nama supplier"],"minimum":["stok minimum","minimum"]}
            idx={}
            for k,als in aliases.items():
                for a in als:
                    if a in headers: idx[k]=headers.index(a)
            if "sku" not in idx or "name" not in idx: raise ValueError("Header wajib: SKU dan Nama Barang")
            count=0
            with db() as c:
                with c.cursor() as cur:
                    for row in ws.iter_rows(min_row=2,values_only=True):
                        sku=str(row[idx["sku"]]).strip() if row[idx["sku"]] is not None else ""
                        name=str(row[idx["name"]]).strip() if row[idx["name"]] is not None else ""
                        if not sku or not name: continue
                        def val(k,default=""):
                            return row[idx[k]] if k in idx and row[idx[k]] is not None else default
                        cur.execute("""INSERT INTO products(sku,name,category,unit,buy_price,supplier,minimum,active)
                          VALUES(%s,%s,%s,%s,%s,%s,%s,TRUE)
                          ON CONFLICT(sku) DO UPDATE SET name=EXCLUDED.name,category=EXCLUDED.category,unit=EXCLUDED.unit,
                          buy_price=EXCLUDED.buy_price,supplier=EXCLUDED.supplier,minimum=EXCLUDED.minimum,active=TRUE""",
                          (sku,name,val("category"),val("unit","PCS"),float(val("buy_price",0) or 0),val("supplier"),float(val("minimum",0) or 0)))
                        count+=1
            log("Upload Barang",f"{count} SKU"); flash(f"{count} barang berhasil disimpan.","ok")
        except Exception as e: flash("Excel gagal diproses: "+str(e),"err")
        finally:
            try:path.unlink()
            except:pass
        return redirect(url_for("products"))
    q=request.args.get("q","").strip()
    with db() as c:
        with c.cursor() as cur:
            cur.execute("SELECT * FROM products WHERE sku ILIKE %s OR name ILIKE %s OR supplier ILIKE %s ORDER BY name",(f"%{q}%",f"%{q}%",f"%{q}%")); rows=cur.fetchall()
    return render_template("products.html", products=rows, q=q)

@app.post("/products/<int:id>/toggle")
def toggle_product(id):
    with db() as c:
        with c.cursor() as cur:
            cur.execute("UPDATE products SET active=NOT active WHERE id=%s RETURNING name,active",(id,)); r=cur.fetchone()
    log("Ubah Status Barang",r["name"]); return redirect(url_for("products"))

@app.route("/inbound",methods=["GET","POST"])
def inbound():
    if request.method=="POST":
        sku=request.form["sku"].strip()
        with db() as c:
            with c.cursor() as cur:
                cur.execute("SELECT * FROM products WHERE sku=%s",(sku,)); p=cur.fetchone()
                if not p: flash("SKU tidak ditemukan.","err"); return redirect(url_for("inbound"))
                qty=float(request.form.get("qty",1) or 1); no=next_no("INB","inbounds")
                cur.execute("""INSERT INTO inbounds(number,arrival_date,supplier_code,supplier_name,sku,exp_date,product_name,price,qty)
                               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (no,request.form["arrival_date"],request.form["supplier_code"],request.form["supplier_name"],sku,request.form["exp_date"],p["name"],float(request.form["price"] or p["buy_price"]),qty))
                cur.execute("UPDATE products SET stock=stock+%s WHERE sku=%s",(qty,sku))
        log("Inbound",f"{no} {sku} +{qty}"); flash(f"Inbound {no} tersimpan.","ok"); return redirect(url_for("inbound"))
    with db() as c:
        with c.cursor() as cur:
            cur.execute("SELECT * FROM inbounds ORDER BY id DESC LIMIT 100"); rows=cur.fetchall()
            cur.execute("SELECT * FROM products WHERE active=TRUE ORDER BY name"); prods=cur.fetchall()
    return render_template("inbound.html",inbounds=rows,products=prods)

@app.route("/requests",methods=["GET","POST"])
def requests():
    if request.method=="POST":
        f=request.files.get("file"); store_id=request.form.get("store_id")
        if not f or not f.filename.lower().endswith((".xlsx",".xlsm")):
            flash("Pilih Excel request.","err"); return redirect(url_for("requests"))
        path=UPLOADS/(str(uuid.uuid4())+"_"+f.filename); f.save(path)
        try:
            ws=load_workbook(path,data_only=True).active; h=[str(x.value).strip().lower() if x.value is not None else "" for x in ws[1]]
            def col(names):
                for n in names:
                    if n in h:return h.index(n)
                return None
            si,ni,qi=col(["sku"]),col(["nama barang","nama","name"]),col(["qty request","qty","quantity"])
            if si is None or qi is None: raise ValueError("Header wajib: SKU dan Qty Request")
            no=next_no("REQ","requests")
            with db() as c:
                with c.cursor() as cur:
                    cur.execute("INSERT INTO requests(number,store_id,request_date) VALUES(%s,%s,CURRENT_DATE) RETURNING id",(no,store_id)); rid=cur.fetchone()["id"]
                    count=0
                    for row in ws.iter_rows(min_row=2,values_only=True):
                        sku=str(row[si]).strip() if row[si] is not None else ""; qty=float(row[qi] or 0)
                        if not sku or qty<=0: continue
                        cur.execute("SELECT name FROM products WHERE sku=%s",(sku,)); p=cur.fetchone()
                        name=str(row[ni]).strip() if ni is not None and row[ni] else (p["name"] if p else "")
                        cur.execute("INSERT INTO request_items(request_id,sku,product_name,qty) VALUES(%s,%s,%s,%s)",(rid,sku,name,qty)); count+=1
            log("Upload Request",f"{no}, {count} item"); flash(f"Request {no} tersimpan ({count} item).","ok")
        except Exception as e: flash("Request gagal diproses: "+str(e),"err")
        finally:
            try:path.unlink()
            except:pass
        return redirect(url_for("requests"))
    with db() as c:
        with c.cursor() as cur:
            cur.execute("""SELECT r.*,s.name AS store_name,(SELECT COUNT(*) FROM request_items i WHERE i.request_id=r.id) AS items
                          FROM requests r LEFT JOIN stores s ON s.id=r.store_id ORDER BY r.id DESC LIMIT 100"""); rows=cur.fetchall()
            cur.execute("SELECT * FROM stores WHERE active=TRUE ORDER BY name"); stores_=cur.fetchall()
    return render_template("requests.html",requests=rows,stores=stores_)

@app.route("/fulfillment")
def fulfillment():
    result=[]
    with db() as c:
        with c.cursor() as cur:
            cur.execute("SELECT r.*,s.name AS store_name FROM requests r LEFT JOIN stores s ON s.id=r.store_id ORDER BY r.id DESC LIMIT 100"); reqs=cur.fetchall()
            for r in reqs:
                cur.execute("SELECT * FROM request_items WHERE request_id=%s",(r["id"],)); items=cur.fetchall()
                for i in items:
                    cur.execute("SELECT stock FROM products WHERE sku=%s",(i["sku"],)); p=cur.fetchone(); stock=float(p["stock"]) if p else 0
                    fulfill=min(float(i["qty"]),stock); shortage=max(float(i["qty"])-stock,0)
                    result.append(dict(number=r["number"],store=r["store_name"],sku=i["sku"],product=i["product_name"],request=float(i["qty"]),stock=stock,fulfill=fulfill,shortage=shortage,status="FULL" if stock>=float(i["qty"]) else ("PARTIAL" if stock>0 else "OOS")))
    return render_template("fulfillment.html",items=result)

@app.route("/settings",methods=["GET","POST"])
def settings():
    if request.method=="POST":
        with db() as c:
            with c.cursor() as cur: cur.execute("UPDATE settings SET language=%s,theme=%s WHERE id=1",(request.form["language"],request.form["theme"]))
        log("Settings","Bahasa/tema diperbarui"); flash("Pengaturan disimpan.","ok"); return redirect(url_for("settings"))
    return render_template("settings.html")

@app.route("/logs")
def logs():
    with db() as c:
        with c.cursor() as cur: cur.execute("SELECT * FROM logs ORDER BY id DESC LIMIT 300"); rows=cur.fetchall()
    return render_template("logs.html",logs=rows)

@app.route("/download/template/products")
def product_template():
    wb=Workbook(); ws=wb.active; ws.append(["SKU","Nama Barang","Kategori","Satuan","Harga Pembelian","Supplier","Stok Minimum"]); ws.append(["SKU001","Contoh Barang","Kategori A","PCS",10000,"Supplier A",5])
    p=UPLOADS/"template_barang.xlsx"; wb.save(p); return send_file(p,as_attachment=True,download_name="template_barang.xlsx")

@app.route("/download/template/requests")
def request_template():
    wb=Workbook(); ws=wb.active; ws.append(["SKU","Nama Barang","Qty Request"]); ws.append(["SKU001","Contoh Barang",10])
    p=UPLOADS/"template_request.xlsx"; wb.save(p); return send_file(p,as_attachment=True,download_name="template_request.xlsx")

@app.get("/health")
def health(): return {"status":"ok"}

if __name__=="__main__":
    init_db()
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)))
